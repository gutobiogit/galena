from openpyxl import load_workbook
from spreadsheet_entity import Spreadsheet


def split_address(address):
    address_split = address.split(",")
    address_split = [x.strip(' ') for x in address_split]
    state = address_split[-1].split(" ")[0]
    zip_code = address_split[-1].split(" ")[1]
    city = address_split[-2]
    number = address_split[0].split(" ")[-1].strip(" ")
    print(number)
    street = " ".join(address_split[0].split(" ")[0:-1])
    return [street, number, city, state, zip_code]


wb = load_workbook('tabela-de-exemplo.xlsx')
ws = wb['results-20220124-165033']
objects = []
max_columns = ws.max_column+2
emails = ws["A4":f"A{max_columns}"]
names = ws["B4":f"B{max_columns}"]
groups = ws["C4":f"C{max_columns}"]
group_names = ws["D4":f"D{max_columns}"]
cpfs = ws["E4":f"E{max_columns}"]
telephones = ws["F4":f"F{max_columns}"]
birthdays = ws["G4":f"G{max_columns}"]
address = ws["H4":f"H{max_columns}"]
for count in range(1, max_columns-3):

    obj = Spreadsheet()
    obj.email = emails[count][0].value
    obj.name = names[count][0].value
    obj.group = groups[count][0].value
    obj.group_name = group_names[count][0].value
    obj.cpf = cpfs[count][0].value
    obj.telephone = telephones[count][0].value
    obj.birthday = birthdays[count][0].value
    obj.address = split_address(address[count][0].value)
    objects.append(obj)
    print(obj.email)
    print(obj.name)
    print(obj.group)
    print(obj.group_name)
    print(obj.cpf)
    print(obj.telephone)
    print(obj.birthday)
    print(obj.address)
    print("\n")
    print("\n")

    objects.append(obj)
