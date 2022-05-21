import ast
from openpyxl import load_workbook
from spreadsheet_entity import Spreadsheet


FILE = 'src/database/tabela-de-exemplo.xlsx'
workbook = load_workbook(FILE)
ws = workbook['results-20220124-165033']

def split_address(address):
    address_split = address.split(",")
    address_split = [x.strip(' ') for x in address_split]
    state = address_split[-1].split(" ")[0]
    zip_code = address_split[-1].split(" ")[1]
    city = address_split[-2]
    number = address_split[0].split(" ")[-1].strip(" ")
    street = " ".join(address_split[0].split(" ")[0:-1])
    return [street, number, city, state, zip_code]

def build_address(address):
    street = address['street']
    number = address['number']
    city = address['city']
    state = address['state']
    zip_code = address['zip_code']
    return f"{street} {number}, {city}, {state} {zip_code}"

def get_one_record(row_id):
    row_id+=4
    obj = Spreadsheet()
    obj.email = ws[f"A{row_id}"].value
    obj.name = ws[f"B{row_id}"].value
    obj.group = ws[f"C{row_id}"].value
    obj.group_name = ws[f"D{row_id}"].value
    obj.cpf = ws[f"E{row_id}"].value
    obj.telephone = ws[f"F{row_id}"].value
    obj.birthday = ws[f"G{row_id}"].value
    obj.address = split_address(ws[f"H{row_id}"].value)
    return obj

def get_all_records():
    objects = []
    max_columns = len([row for row in ws if not all([cell.value is None for cell in row])])+2
    emails = ws["A4":f"A{max_columns}"]
    names = ws["B4":f"B{max_columns}"]
    groups = ws["C4":f"C{max_columns}"]
    group_names = ws["D4":f"D{max_columns}"]
    cpfs = ws["E4":f"E{max_columns}"]
    telephones = ws["F4":f"F{max_columns}"]
    birthdays = ws["G4":f"G{max_columns}"]
    address = ws["H4":f"H{max_columns}"]
    for count in range(1, max_columns-3):
        obj = Spreadsheet()
        obj.email = emails[count][0].value
        obj.name = names[count][0].value
        obj.group = groups[count][0].value
        obj.group_name = group_names[count][0].value
        obj.cpf = cpfs[count][0].value
        obj.telephone = telephones[count][0].value
        obj.birthday = birthdays[count][0].value
        obj.address = split_address(address[count][0].value)
        objects.append(obj)
    return objects

def change_row(change_id, row):
    row_id=change_id+4
    ws[f"A{row_id}"] = row.email.email_address
    ws[f"B{row_id}"] = row.name
    ws[f"C{row_id}"] = row.group
    ws[f"D{row_id}"] = row.group_name
    ws[f"E{row_id}"] = ast.literal_eval(row.cpf)['cpf']
    ws[f"F{row_id}"] = ast.literal_eval(row.telephone)['telephone']
    ws[f"G{row_id}"] = ast.literal_eval(row.birthday)['birthday']
    ws[f"H{row_id}"] = build_address(ast.literal_eval(row.address))
    workbook.save(FILE)
    return f"Galerner alterado na posição {change_id}"

def add_row(row):
    insert_row = len([row for row in ws if not all([cell.value is None for cell in row])])+3
    ws.insert_rows(insert_row)
    ws[f"A{insert_row}"] = row.email.email_address
    ws[f"B{insert_row}"] = row.name
    ws[f"C{insert_row}"] = row.group
    ws[f"D{insert_row}"] = row.group_name
    ws[f"E{insert_row}"] = ast.literal_eval(row.cpf)['cpf']
    ws[f"F{insert_row}"] = ast.literal_eval(row.telephone)['telephone']
    ws[f"G{insert_row}"] = ast.literal_eval(row.birthday)['birthday']
    ws[f"H{insert_row}"] = build_address(ast.literal_eval(row.address))
    workbook.save(FILE)
    return "Galerner inserido na posição {}".format(insert_row)

def delete_row(row_number):
    ws.delete_rows(row_number+4)
    workbook.save(FILE)
    return f"Galerner deletado na posição {row_number}"
